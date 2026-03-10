# import streamlit as st
# import pandas as pd
# import re

# st.set_page_config(layout="wide")

# st.title("Order Message Generator")

# uploaded_file = st.file_uploader("Upload TableInputDaily.xlsx", type=["xlsx"])

# # --- Helpers ---
# def format_issue_size(size):
#     try:
#         size = float(size)
#         if size >= 1000:
#             return f"{size/1000:g}bn"
#         else:
#             return f"{int(size)}"
#     except:
#         return ""

# def parse_order(order):
#     if pd.isna(order):
#         return None, None
    
#     match = re.match(r"(\d+)\((.*?)\)", str(order))
#     if match:
#         return match.group(1), match.group(2)
#     return None, None

# # --- Main ---
# if uploaded_file:

#     df = pd.read_excel(uploaded_file)

#     st.subheader("Deal Table")

#     edited_df = st.data_editor(
#         df,
#         use_container_width=True,
#         num_rows="dynamic"
#     )

#     st.markdown("---")

#     if st.button("Generate Messages"):

#         grouped = edited_df.groupby("Issuer")
#         messages = []

#         for issuer, group in grouped:

#             lines = []
#             issue_sizes = []

#             for _, row in group.iterrows():

#                 order_x, order_y = parse_order(row["Order"])
#                 if not order_x:
#                     continue

#                 issue_size = format_issue_size(row["Issue Size"])
#                 issue_sizes.append(issue_size)

#                 line = f'{row["CCY"]} {row["Tenor"]}: {order_x}mm -> min alloc {order_y}'
#                 lines.append(line)

#             if not lines:
#                 continue

#             if len(set(issue_sizes)) == 1:
#                 assumption = f"(assumes {issue_sizes[0]} mm per tranche)"
#             else:
#                 assumption = "\n".join([f"(assumes {s} mm tranche)" for s in issue_sizes])

#             msg = f"""Brevan Howard Gups Abu Dhabi
# {issuer}
# Order post prior deal support / reverses

# {chr(10).join(lines)}
# {assumption}

# Outright
# Thx
# Order subject to final size and price
# Order also in direct books (not a dupe)
# """
#             messages.append(msg)

#         st.subheader("Generated Messages")

#         full_output = "\n\n--------------------------\n\n".join(messages)

#         st.text_area("Messages", full_output, height=400)

#         st.download_button(
#             "Download Messages",
#             full_output,
#             file_name="order_messages.txt"
#         )









# import streamlit as st
# import pandas as pd
# import re
# from openpyxl import load_workbook
# from tempfile import NamedTemporaryFile

# st.set_page_config(layout="wide")

# st.title("Order Message Generator")

# # --- Helpers ---
# def format_issue_size(size):
#     try:
#         size = float(size)
#         if size >= 1000:
#             return f"{size/1000:g}bn"
#         else:
#             return f"{int(size)}"
#     except:
#         return ""

# def parse_order(order):
#     if pd.isna(order):
#         return None, None
    
#     match = re.match(r"(\d+)\((.*?)\)", str(order))
#     if match:
#         return match.group(1), match.group(2)
#     return None, None

# # --- Upload Excel ---
# uploaded_file = st.file_uploader("Upload Excel", type=["xlsx","xlsm"])

# if uploaded_file:

#     # Load workbook FULLY (keeps formulas & tables)
#     temp = NamedTemporaryFile(delete=False, suffix='.xlsx')
#     temp.write(uploaded_file.getvalue())
#     temp.close()

#     wb = load_workbook(temp.name, data_only=False)

#     ws = wb.active   # <- NO sheet name dependency

#     data = ws.values
#     cols = next(data)
#     df = pd.DataFrame(data, columns=cols)


#     edited_df = st.data_editor(
#         df,
#         use_container_width=True,
#         num_rows="dynamic"
#     )


#     # ---------------- LIVE CALCULATIONS ---------------- #

#     def extract_order_parts(order):
#         if pd.isna(order):
#             return None, None
#         m = re.match(r"(\d+)\((.*?)\)", str(order))
#         if m:
#             try:
#                 return float(m.group(1)), m.group(2)
#             except:
#                 return None, None
#         return None, None


#     calc_df = edited_df.copy()


#     # Order Multiple
#     if "Order Multiple" in calc_df.columns:
#         def calc_order_multiple(row):
#             x, y = extract_order_parts(row.get("Order"))
#             try:
#                 if x and y and "%" not in str(y):
#                     return round(x / float(y), 1)
#             except:
#                 return None
#             return None

#         calc_df["Order Multiple"] = calc_df.apply(calc_order_multiple, axis=1)


#     # Book multiple
#     if "Book multiple" in calc_df.columns:
#         calc_df["Book multiple"] = (
#             pd.to_numeric(calc_df.get("Book size"), errors="coerce") /
#             pd.to_numeric(calc_df.get("Issue Size"), errors="coerce")
#         )


#     # pnl (ccy)
#     if "pnl (ccy)" in calc_df.columns:
#         calc_df["pnl (ccy)"] = (
#             pd.to_numeric(calc_df.get("Sold"), errors="coerce") * 1_000_000 *
#             pd.to_numeric(calc_df.get("Margin (cents)"), errors="coerce") / 10000
#         )


#     # pnl (usd)
#     if "pnl(usd)" in calc_df.columns:
#         def calc_usd(row):
#             pnl = row.get("pnl (ccy)")
#             ccy = row.get("CCY")

#             if pd.isna(pnl):
#                 return None

#             if ccy == "EUR":
#                 return pnl * 1.18
#             elif ccy == "GBP":
#                 return pnl * 1.35
#             return pnl

#         calc_df["pnl(usd)"] = calc_df.apply(calc_usd, axis=1)


#     # alloc as % of issue
#     if "alloc as a % of issue" in calc_df.columns:
#         calc_df["alloc as a % of issue"] = (
#             pd.to_numeric(calc_df.get("Allocated"), errors="coerce") /
#             pd.to_numeric(calc_df.get("Issue Size"), errors="coerce")
#         )


#     # alloc as % of order
#     if "alloc as a % of order" in calc_df.columns:
#         calc_df["alloc as a % of order"] = (
#             pd.to_numeric(calc_df.get("Allocated"), errors="coerce") /
#             pd.to_numeric(calc_df.get("Desired Allocation"), errors="coerce")
#         )


#     if "Select" not in df.columns:
#         df.insert(0, "Select", False)

#     st.subheader("Deal Table")

#     edited_df = st.data_editor(
#         df,
#         use_container_width=True,
#         num_rows="dynamic"
#     )

#     col1, col2 = st.columns(2)

#     # --- SAFE SAVE BUTTON ---
#     with col1:
#         if st.button("💾 Save Changes (Safe)"):

#             edited = edited_df.drop(columns=["Select"], errors="ignore")

#             # Write ONLY cell values back
#             for r_idx, row in edited.iterrows():
#                 for c_idx, value in enumerate(row):

#                     # FIX: Convert pandas NA to real empty Excel cell
#                     if pd.isna(value):
#                         value = None

#                     ws.cell(row=r_idx+2, column=c_idx+1).value = value

#             wb.save(temp.name)

#             with open(temp.name, "rb") as f:
#                 st.download_button(
#                     "⬇ Download Updated Excel",
#                     f,
#                     file_name=uploaded_file.name
#                 )

#             st.success("Saved safely. Tables, formulas & sheets preserved.")

#     # --- GENERATE SELECTED ISSUER ---
#     with col2:
#         if st.button("📨 Generate Message for Selected Issuer"):

#             selected_rows = edited_df[edited_df["Select"] == True]

#             if selected_rows.empty:
#                 st.warning("Please tick a row first.")
#             else:
#                 issuer = selected_rows.iloc[0]["Issuer"]

#                 issuer_rows = edited_df[edited_df["Issuer"] == issuer]

#                 lines = []
#                 issue_sizes = []

#                 for _, row in issuer_rows.iterrows():

#                     order_x, order_y = parse_order(row["Order"])
#                     if not order_x:
#                         continue

#                     issue_size = format_issue_size(row["Issue Size"])
#                     issue_sizes.append(issue_size)

#                     line = f'{row["CCY"]} {row["Tenor"]}: {order_x}mm -> min alloc {order_y}'
#                     lines.append(line)

#                 if not lines:
#                     st.warning("No valid orders for this issuer.")
#                 else:

#                     if len(set(issue_sizes)) == 1:
#                         assumption = f"(assumes {issue_sizes[0]} mm per tranche)"
#                     else:
#                         assumption = "\n".join([f"(assumes {s} mm tranche)" for s in issue_sizes])

#                     msg = f"""Brevan Howard Gups Abu Dhabi
# {issuer}
# Order post prior deal support / reverses

# {chr(10).join(lines)}
# {assumption}

# Outright
# Thx
# Order subject to final size and price
# Order also in direct books (not a dupe)
# """

#                     st.subheader("Generated Message")
#                     st.text_area("Message", msg, height=300)









#-----------------------------------------------------------------------------------------
# import streamlit as st
# import pandas as pd
# import re
# from openpyxl import load_workbook
# from tempfile import NamedTemporaryFile
# import numpy as np


# st.set_page_config(layout="wide")

# st.title("Order Message Generator")

# # ---------------- PARSERS ----------------

# def parse_order(order):
#     if pd.isna(order):
#         return None, None

#     match = re.search(r'([\d\.]+)\(([^)]+)\)', str(order))
#     if not match:
#         return None, None

#     # If bracket contains %, skip calculation
#     if "%" in match.group(2):
#         return None, None

#     try:
#         return float(match.group(1)), float(match.group(2))
#     except:
#         return None, None

# def format_issue_size(size):
#     try:
#         size = float(size)
#         if size >= 1000:
#             return f"{size/1000:g}bn"
#         else:
#             return f"{int(size)}"
#     except:
#         return ""

# # ---------------- CALCULATIONS ----------------

# def calculate_fields(df):

#     def safe_div(a,b):
#         try:
#             if b in [0,None,np.nan]:
#                 return None
#             return a/b
#         except:
#             return None

#     order_multiple = []
#     book_multiple = []
#     pnl_ccy = []
#     pnl_usd = []
#     alloc_issue = []
#     alloc_order = []

#     for _, row in df.iterrows():

#         # ORDER MULTIPLE
#         o1, o2 = parse_order(row.get("Order"))
#         if o1 and o2:
#             order_multiple.append(round(o1/o2,1))
#         else:
#             order_multiple.append(None)

#         # BOOK MULTIPLE
#         try:
#             val = safe_div(row.get("Book size"), row.get("Issue Size"))
#             book_multiple.append(round(val,1) if val else None)
#         except:
#             book_multiple.append(None)

#         # PNL CCY
#         try:
#             pnl = row.get("Sold") * 1000000 * row.get("Margin (cents)") / 10000
#             pnl_ccy.append(pnl)
#         except:
#             pnl_ccy.append(None)

#         # PNL USD
#         try:
#             ccy = row.get("CCY")
#             base = pnl_ccy[-1]

#             if base is None:
#                 pnl_usd.append(None)
#             elif ccy == "EUR":
#                 pnl_usd.append(base * 1.18)
#             elif ccy == "GBP":
#                 pnl_usd.append(base * 1.35)
#             else:
#                 pnl_usd.append(base)
#         except:
#             pnl_usd.append(None)

#         # ALLOC % ISSUE
#         try:
#             val = safe_div(row.get("Allocated"), row.get("Issue Size"))
#             alloc_issue.append(val)
#         except:
#             alloc_issue.append(None)

#         # ALLOC % ORDER
#         try:
#             val = safe_div(row.get("Allocated"), row.get("Desired Allocation"))
#             alloc_order.append(val)
#         except:
#             alloc_order.append(None)

#     df["Order multiple"] = order_multiple
#     df["Book multiple"] = book_multiple
#     df["pnl (ccy)"] = pnl_ccy
#     df["pnl (usd)"] = pnl_usd
#     df["alloc as a % of issue"] = alloc_issue
#     df["alloc as a % of order"] = alloc_order

#     return df

# # ---------------- UPLOAD ----------------

# uploaded_file = st.file_uploader("Upload Excel", type=["xlsx","xlsm"])

# if uploaded_file:

#     temp = NamedTemporaryFile(delete=False, suffix='.xlsx')
#     temp.write(uploaded_file.getvalue())
#     temp.close()

#     wb = load_workbook(temp.name, data_only=False)
#     ws = wb.active

#     data = ws.values
#     cols = next(data)
#     df = pd.DataFrame(data, columns=cols)

#     if "Select" not in df.columns:
#         df.insert(0, "Select", False)

#     df = calculate_fields(df)

#     st.subheader("Deal Table")

#     edited_df = st.data_editor(
#         df,
#         use_container_width=True,
#         num_rows="dynamic",
#         disabled=[
#             "Order multiple",
#             "Book multiple",
#             "pnl (ccy)",
#             "pnl (usd)",
#             "alloc as % of issue",
#             "alloc as % of order"
#         ]
#     )

#     edited_df = calculate_fields(edited_df)

#     col1, col2 = st.columns(2)

#     # ---------------- SAVE ----------------

#     with col1:
#         if st.button("💾 Save Changes (Safe)"):
#             edited = edited_df.drop(columns=["Select"], errors="ignore")

#             for r_idx, row in edited.iterrows():
#                 for c_idx, value in enumerate(row):
#                     if pd.isna(value):
#                         value = None
#                     ws.cell(row=r_idx+2, column=c_idx+1).value = value

#             wb.save(temp.name)

#             with open(temp.name, "rb") as f:
#                 st.download_button(
#                     "⬇ Download Updated Excel",
#                     f,
#                     file_name=uploaded_file.name
#                 )

#             st.success("Saved safely.")

#     # ---------------- MESSAGE ----------------

#     with col2:
#         if st.button("📨 Generate Message for Selected Issuer"):

#             selected_rows = edited_df[edited_df["Select"] == True]

#             if selected_rows.empty:
#                 st.warning("Please tick a row first.")
#             else:
#                 issuer = selected_rows.iloc[0]["Issuer"]
#                 issuer_rows = edited_df[edited_df["Issuer"] == issuer]

#                 lines = []
#                 issue_sizes = []

#                 for _, row in issuer_rows.iterrows():

#                     o1, o2 = parse_order(row["Order"])
#                     if not o1:
#                         continue

#                     issue_size = format_issue_size(row["Issue Size"])
#                     issue_sizes.append(issue_size)

#                     line = f'{row["CCY"]} {row["Tenor"]}: {int(o1)}mm -> min alloc {int(o2)}'
#                     lines.append(line)

#                 if not lines:
#                     st.warning("No valid orders.")
#                 else:

#                     if len(set(issue_sizes)) == 1:
#                         assumption = f"(assumes {issue_sizes[0]} mm per tranche)"
#                     else:
#                         assumption = "\n".join([f"(assumes {s} mm tranche)" for s in issue_sizes])

#                     msg = f"""Brevan Howard Gups Abu Dhabi
# {issuer}
# Order post prior deal support / reverses

# {chr(10).join(lines)}
# {assumption}

# Outright
# Thx
# Order subject to final size and price
# Order also in direct books (not a dupe)
# """

#                     st.subheader("Generated Message")
#                     st.text_area("Message", msg, height=300)


import streamlit as st
import pandas as pd
import re
import os
from tempfile import NamedTemporaryFile
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

st.set_page_config(layout="wide", page_title="Deal Table")
st.title("Deal Table")

# ─── Columns the app calculates. Everything else is user-editable. ────────────
CALC_COLS = [
    "Order Multiple",
    "Book multiple",
    "pnl (ccy)",
    "pnl(usd)",
    "alloc as a % of issue",
    "alloc as a % of order",
]

# ─── CALCULATIONS ─────────────────────────────────────────────────────────────

def sf(v):
    """Safe string → float. Returns None on any failure."""
    if v is None:
        return None
    s = str(v).strip().replace(",", "").replace("%", "")
    if s in ("", "None", "nan", "NaN", "#VALUE!", "#DIV/0!"):
        return None
    try:
        return float(s)
    except Exception:
        return None

def calc_order_multiple(order):
    """60(20) → 3.0x  (outside ÷ inside).  % in bracket → None."""
    if not order:
        return None
    m = re.match(r"^(\d+(?:\.\d+)?)\((\d+(?:\.\d+)?)\)\s*$", str(order).strip())
    if m:
        outside, inside = float(m.group(1)), float(m.group(2))
        return f"{outside/inside:.1f}x" if inside != 0 else None
    return None

def calc_book_multiple(book_size, issue_size):
    b, i = sf(book_size), sf(issue_size)
    if b is not None and i and i != 0:
        return f"{b/i:.1f}x"
    return None

def calc_pnl_ccy(sold, margin):
    s, m = sf(sold), sf(margin)
    if s is not None and m is not None:
        return round(s * 1_000_000 * m / 10_000, 0)
    return None

def calc_pnl_usd(pnl_c, ccy):
    if pnl_c is None:
        return None
    c = str(ccy or "").strip().upper()
    if c == "EUR":
        return round(pnl_c * 1.18, 0)
    if c == "GBP":
        return round(pnl_c * 1.35, 0)
    return pnl_c

def calc_alloc_pct_issue(allocated, issue_size):
    a, i = sf(allocated), sf(issue_size)
    if a is not None and i and i != 0:
        return round(a / i * 100, 2)
    return None

def calc_alloc_pct_order(allocated, desired):
    a, d = sf(allocated), sf(desired)
    if a is not None and d and d != 0:
        return round(a / d * 100, 2)
    return None

def recompute(df: pd.DataFrame) -> pd.DataFrame:
    """Recompute all 6 calc columns from input values. Returns updated dataframe."""
    df = df.copy()
    # Ensure calc cols exist and are object dtype so we can write any value type
    for c in CALC_COLS:
        if c not in df.columns:
            df[c] = None
        df[c] = df[c].astype(object)

    for i, row in df.iterrows():
        pc = calc_pnl_ccy(row.get("Sold"), row.get("Margin (cents)"))
        a  = sf(row.get("Allocated"))
        i2 = sf(row.get("Issue Size"))
        d  = sf(row.get("Desired Allocation"))
        df.at[i, "Order Multiple"]          = calc_order_multiple(row.get("Order"))
        df.at[i, "Book multiple"]            = calc_book_multiple(row.get("Book size"), row.get("Issue Size"))
        df.at[i, "pnl (ccy)"]               = pc
        df.at[i, "pnl(usd)"]                = calc_pnl_usd(pc, row.get("CCY"))
        df.at[i, "alloc as a % of issue"]    = round(a/i2*100, 2) if (a is not None and i2 and i2 != 0) else None
        df.at[i, "alloc as a % of order"]    = round(a/d*100,  2) if (a is not None and d  and d  != 0) else None
    return df

# ─── HELPERS ──────────────────────────────────────────────────────────────────

def clean_cell(v):
    """Strip Excel formulas/errors. Return plain string for all values."""
    if v is None:
        return ""
    s = str(v)
    if s.startswith("=") or s.startswith("#"):
        return ""
    return s.strip()

def to_excel_val(v):
    if v is None:
        return None
    if isinstance(v, float) and pd.isna(v):
        return None
    if isinstance(v, str):
        s = v.strip()
        # Try converting numeric strings back to numbers for Excel
        if s == "":
            return None
        try:
            if "." in s:
                return float(s)
            return int(s)
        except Exception:
            return s
    return v

def fix_table_refs(ws, new_row_count: int):
    """
    After inserting/deleting rows, update all Excel table refs so the table
    covers exactly header + data rows. Without this, Excel flags the file corrupt.
    """
    for tname in list(ws.tables.keys()):
        tbl = ws.tables[tname]
        m = re.match(r"([A-Za-z]+)(\d+):([A-Za-z]+)(\d+)", tbl.ref)
        if m:
            tbl.ref = f"{m.group(1)}{m.group(2)}:{m.group(3)}{new_row_count + 1}"

# ─── SESSION STATE ────────────────────────────────────────────────────────────
for k, d in [("temp_path", None), ("fname", None), ("fhash", None), ("df", None)]:
    if k not in st.session_state:
        st.session_state[k] = d

# ─── FILE UPLOAD ──────────────────────────────────────────────────────────────
uploaded = st.file_uploader("Upload your Excel file", type=["xlsx", "xlsm"])

if uploaded:
    h = hash(uploaded.getvalue())
    if st.session_state.fhash != h:
        ext = ".xlsm" if uploaded.name.endswith(".xlsm") else ".xlsx"
        tmp = NamedTemporaryFile(delete=False, suffix=ext)
        tmp.write(uploaded.getvalue())
        tmp.close()
        st.session_state.temp_path = tmp.name
        st.session_state.fname     = uploaded.name
        st.session_state.fhash     = h
        st.session_state.df        = None  # force reload

temp_path = st.session_state.temp_path

# ─── MAIN ─────────────────────────────────────────────────────────────────────
if temp_path and os.path.exists(temp_path):

    # Load workbook to get headers and row count
    wb = load_workbook(temp_path, keep_vba=False, data_only=False)
    ws = wb.worksheets[0]
    all_rows = list(ws.iter_rows())

    if not all_rows:
        st.error("Sheet is empty.")
        st.stop()

    # ── Headers ───────────────────────────────────────────────────────────────
    headers = []
    for c in all_rows[0]:
        h = c.value
        headers.append(str(h).strip("_").strip() if h else f"Col{len(headers)+1}")
    n_cols = len(headers)

    # ── Load data from file into session state (only on first load) ───────────
    if st.session_state.df is None:
        records = []
        for row_cells in all_rows[1:]:
            row = {}
            for ci, h in enumerate(headers):
                v = row_cells[ci].value if ci < len(row_cells) else None
                row[h] = clean_cell(v)
            records.append(row)

        df = pd.DataFrame(records, columns=headers) if records else pd.DataFrame(columns=headers)

        # Drop fully-empty trailing rows
        df = df[df.apply(
            lambda r: any(str(v).strip() not in ("", "None", "nan") for v in r.values), axis=1
        )].reset_index(drop=True)

        # All columns stored as strings — prevents Streamlit type-locking
        for col in df.columns:
            df[col] = df[col].apply(
                lambda x: "" if str(x).strip() in ("None", "nan", "NaN") else str(x).strip()
            )

        # Compute calc cols from loaded data
        df = recompute(df)
        st.session_state.df = df

    df = st.session_state.df.copy()

    # Add select column
    if "_select" not in df.columns:
        df.insert(0, "_select", False)

    # ── Column config: calc cols hidden in top table, everything else editable ──
    col_cfg = {
        "_select": st.column_config.CheckboxColumn("Select", default=False, width="small"),
    }
    for h in headers:
        if h in CALC_COLS:
            # Hidden from top table — shown in bottom summary table instead
            col_cfg[h] = st.column_config.TextColumn(h, disabled=True)
        elif h == "PowerAppsId":
            col_cfg[h] = st.column_config.TextColumn(h, disabled=True)
        else:
            col_cfg[h] = st.column_config.TextColumn(h, disabled=False)

    # Columns to hide from top table (calc cols hidden — shown below instead)
    hidden_in_top = set(CALC_COLS)

    st.caption(f"Sheet: **{ws.title}** | All sheets: {', '.join(wb.sheetnames)}")
    st.caption("Calc columns (Order Multiple, Book multiple, pnl, alloc %) are shown in the summary table below.")

    # ── Single editable table (calc cols hidden) ──────────────────────────────
    top_col_order = [c for c in df.columns if c not in hidden_in_top]
    edited = st.data_editor(
        df[top_col_order],
        column_config=col_cfg,
        use_container_width=True,
        num_rows="dynamic",
        hide_index=True,
        key="tbl",
    )

    # ── Recompute: merge edited input cols back into full df, then recompute ────
    # edited only has visible columns (calc cols are hidden); we need the full df
    # so recompute can read all input columns (Sold, Margin, etc.)
    edited_full = df.copy()
    for col in edited.columns:
        if col in edited_full.columns and col != "_select":
            if len(edited) == len(edited_full):
                edited_full[col] = edited[col].values
            # if rows were added/deleted, rebuild from scratch
    if len(edited) != len(edited_full):
        # Row count changed — use edited as the new base (calc cols will be recomputed)
        edited_full = edited.copy()
        for c in CALC_COLS:
            if c not in edited_full.columns:
                edited_full[c] = None
    refreshed = recompute(edited_full)
    st.session_state.df = refreshed.drop(columns=["_select"], errors="ignore")

    # ── Bottom summary table: BillAndDelivery, CCY, Issue Size + all calc cols ─
    st.subheader("Calculated Columns")
    bottom_cols = []
    for c in ["Issuer", "BillAndDelivery", "CCY", "Issue Size"] + CALC_COLS:
        if c in refreshed.columns:
            bottom_cols.append(c)
    st.dataframe(
        refreshed[bottom_cols],
        use_container_width=True,
        hide_index=True,
    )

    st.divider()
    c1, c2, c3 = st.columns(3)

    # ── SAVE ──────────────────────────────────────────────────────────────────
    with c1:
        if st.button("💾 Save Changes (Safe)", type="primary"):
            try:
                wb2 = load_workbook(temp_path, keep_vba=False, data_only=False)
                ws2 = wb2.worksheets[0]   # only ever touch sheet[0]

                data_to_write = refreshed.drop(columns=["_select"], errors="ignore")
                new_n      = len(data_to_write)
                old_n      = len(all_rows) - 1
                FIRST_DATA = 2
                added      = new_n - old_n

                if added > 0:
                    insert_at = FIRST_DATA + old_n
                    ws2.insert_rows(insert_at, amount=added)
                    # Copy style from last data row into new rows
                    template = insert_at - 1
                    for new_r in range(insert_at, insert_at + added):
                        for ci in range(n_cols):
                            col_letter = get_column_letter(ci + 1)
                            src = ws2[f"{col_letter}{template}"]
                            dst = ws2[f"{col_letter}{new_r}"]
                            if src.has_style:
                                from copy import copy
                                dst.font          = copy(src.font)
                                dst.fill          = copy(src.fill)
                                dst.border        = copy(src.border)
                                dst.alignment     = copy(src.alignment)
                                dst.number_format = src.number_format

                elif added < 0:
                    ws2.delete_rows(FIRST_DATA + new_n, amount=-added)

                # Write all values
                for df_i in range(new_n):
                    ws_row = FIRST_DATA + df_i
                    for ci, col_name in enumerate(headers):
                        if col_name in data_to_write.columns:
                            ws2.cell(row=ws_row, column=ci + 1).value = to_excel_val(
                                data_to_write.iloc[df_i][col_name]
                            )

                # ── CRITICAL: fix Excel table refs to match new row count ────
                # Without this, Excel shows "file is corrupt" when rows are added/removed
                fix_table_refs(ws2, new_n)

                wb2.save(temp_path)
                st.success("✅ Saved. Click Download to get your updated file.")

                with open(temp_path, "rb") as f:
                    st.download_button(
                        "⬇️ Download Updated Excel",
                        data=f.read(),
                        file_name=st.session_state.fname,
                        mime=(
                            "application/vnd.ms-excel.sheet.macroEnabled.12"
                            if st.session_state.fname.endswith(".xlsm")
                            else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        ),
                    )
            except Exception as e:
                st.error(f"❌ Save error: {e}")
                raise

    # ── ORDER MESSAGE ─────────────────────────────────────────────────────────
    with c2:
        if st.button("📨 Generate Message for Selected Issuer"):
            selected = edited[edited["_select"] == True]

            if selected.empty:
                st.warning("Tick at least one row first.")
            else:
                cols_lower = {c.lower().strip(): c for c in edited.columns}
                issuer_col = cols_lower.get("issuer") or cols_lower.get("billanddelivery")
                ccy_col    = cols_lower.get("ccy")
                tenor_col  = cols_lower.get("tenor")
                order_col  = cols_lower.get("order")
                issue_col = cols_lower.get("issue size")

                if not issuer_col:
                    st.error("Cannot find Issuer column.")
                elif not order_col:
                    st.error("Cannot find Order column.")
                else:
                    issuer = str(selected.iloc[0][issuer_col]).strip()
                    issuer_rows = edited[edited[issuer_col].astype(str).str.strip() == issuer]

                    lines = []
                    for _, row in issuer_rows.iterrows():
                        order_val = str(row.get(order_col, "")).strip()
                        if not order_val or order_val in ("None", "nan", ""):
                            continue
                        m_size = re.match(r"^(\d+(?:\.\d+)?)", order_val)
                        m_alloc = re.match(r"^\d+(?:\.\d+)?\((.*?)\)$", order_val)
                        if not m_size:
                            continue
                        order_size  = int(float(m_size.group(1)))
                        issue_size = str(row.get(issue_col, "")).strip() if issue_col else ""
                        order_alloc = m_alloc.group(1) if m_alloc else None
                        ccy   = str(row.get(ccy_col,   "")).strip() if ccy_col   else ""
                        tenor = str(row.get(tenor_col, "")).strip() if tenor_col else ""
                        if order_alloc:
                            alloc_str = f" min alloc {order_alloc}mm" if not order_alloc.strip().endswith("%") else f" min alloc {order_alloc}"
                        else:
                            alloc_str = ""
                        #alloc_str = f" min alloc {order_alloc}" if order_alloc else ""
                        lines.append(f"{ccy} {tenor} --> {order_size}mm{alloc_str}\n (assume {issue_size}mm issue)")
                        #lines.append(f"{ccy} {tenor} --> {order_size}mm{alloc_str}")

                    if not lines:
                        st.warning("No valid Orders found for this issuer.")
                    else:
                        msg = (
                            f"Brevan Howard Gups Abu Dhabi\n"
                            f"{issuer}\n"
                            f"Order post prior deal supports/reverses/ioi\n\n"
                            + "\n".join(lines)
                            + "\n\nOutright\nThx\n\n"
                            "Order subject to final size and price\n"
                            "Order also in direct books (not a dupe)"
                        )
                        st.subheader("Generated Message")
                        st.text_area("", value=msg, height=350, key="msg_out")
                        st.caption("Select all (Ctrl+A / Cmd+A) then copy.")
    # ── INVESTOR EMAIL ────────────────────────────────────────────────────────
    with c3:
        if st.button("📧 Generate Investor Email"):
            selected = edited[edited["_select"] == True]

            if selected.empty:
                st.warning("Tick at least one row first.")
            else:
                cols_lower = {c.lower().strip(): c for c in edited.columns}
                issuer_col = cols_lower.get("issuer") or cols_lower.get("billanddelivery")
                order_col  = cols_lower.get("order")
                issue_col  = cols_lower.get("issue size")

                if not issuer_col:
                    st.error("Cannot find Issuer column.")
                elif not order_col:
                    st.error("Cannot find Order column.")
                else:
                    issuer = str(selected.iloc[0][issuer_col]).strip()
                    issuer_rows = edited[edited[issuer_col].astype(str).str.strip() == issuer]

                    tranche_pcts = []
                    total_alloc  = 0.0
                    total_issue  = 0.0

                    for _, row in issuer_rows.iterrows():
                        order_val = str(row.get(order_col, "")).strip()
                        issue_val = str(row.get(issue_col, "")).strip() if issue_col else ""
                        m_alloc   = re.match(r"^\d+(?:\.\d+)?\((\d+(?:\.\d+)?)\)$", order_val)
                        i         = sf(issue_val)
                        if m_alloc and i and i != 0:
                            alloc = float(m_alloc.group(1))
                            tranche_pcts.append(round(alloc / i * 100))
                            total_alloc += alloc
                            total_issue += i

                    if not tranche_pcts:
                        st.warning("No valid Order(min alloc) and Issue Size found for this issuer.")
                    else:
                        # If all tranches have the same %, say "X% each tranche"
                        # Otherwise use combined total alloc / total issue
                        if len(set(tranche_pcts)) == 1 and len(tranche_pcts) > 1:
                            pct_str = f"{tranche_pcts[0]}% each tranche"
                        else:
                            combined_pct = round(total_alloc / total_issue * 100)
                            pct_str = f"{combined_pct}%"

                        email = (
                            f"BREVAN HOWARD GUPS ABU DHABI\n\n"
                            f"Dear Team {issuer},\n\n"
                            f"Great to see a new deal go live today!\n\n"
                            f"FYI we have let the lead banks know we are targeting an allocation "
                            f"of the new deal at {pct_str} of deal size - good into final price "
                            f"and final issue size!\n\n"
                            f"We hope to further grow our partnership with you across currencies and tenors\n\n"
                            f"If you ever visit Abu Dhabi please let us know, we would be delighted "
                            f"to host you at our offices\n\n"
                            f"Best wishes,\n\n"
                            f"~David and Mitesh"
                        )

                        st.subheader("Investor Email")
                        st.text_area("", value=email, height=400, key="email_out")
                        st.caption("Select all (Ctrl+A / Cmd+A) then copy.")
else:
    st.info("Upload your Excel file above to begin")